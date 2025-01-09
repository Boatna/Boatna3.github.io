import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from PIL import Image, ImageTk

# ฟังก์ชันสำหรับค้นหาข้อมูลใน Excel
def search_data():
    query = entry_query.get().strip()
    if not query:
        messagebox.showwarning("คำเตือน", "กรุณากรอกข้อมูลที่ต้องการค้นหา")
        return

    try:
        # โหลดไฟล์ Excel
        workbook = load_workbook("employee_data.xlsx")
        sheet = workbook.active

        # ค้นหาข้อมูล
        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # เริ่มจากแถวที่ 2 เพื่อข้ามหัวตาราง
            if any(query in str(cell) for cell in row):
                results.append(f"รหัสพนักงาน: {row[0]}, ชื่อ-นามสกุล: {row[1]}, แผนก: {row[2]}, โต๊ะที่นั่ง: {row[3]}")

        # แสดงผลการค้นหา
        if results:
            result_text = "\n".join(results)
        else:
            result_text = f"ไม่พบข้อมูลที่ตรงกับ '{query}'"

        # แสดงผลใน Text widget
        text_result.delete(1.0, tk.END)
        text_result.insert(tk.END, result_text)
    except FileNotFoundError:
        messagebox.showerror("ข้อผิดพลาด", "ไม่พบไฟล์ employee_data.xlsx กรุณาเพิ่มไฟล์ในโฟลเดอร์เดียวกันกับโปรแกรม")
    except Exception as e:
        messagebox.showerror("ข้อผิดพลาด", f"เกิดข้อผิดพลาด: {e}")

# ฟังก์ชันสำหรับโหลดรูปภาพอย่างปลอดภัย
def load_image(path, size):
    try:
        image = Image.open(path)
        image = image.resize(size, Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(image)
    except FileNotFoundError:
        messagebox.showerror("ข้อผิดพลาด", f"ไม่พบไฟล์รูปภาพ: {path}")
        return None

# สร้าง UI ด้วย tkinter
root = tk.Tk()
root.title("ระบบค้นหาข้อมูลพนักงาน")

# กำหนดขนาดหน้าต่าง
window_width = 1000
window_height = 700
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# คำนวณตำแหน่งหน้าต่างให้อยู่ตรงกลาง
x_position = (screen_width // 2) - (window_width // 2)
y_position = (screen_height // 2) - (window_height // 2)
root.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
root.config(bg="#f4faff")

# โลโก้ 1
logo1_photo = load_image("Bitwise-BW.png", (50, 50))
if logo1_photo:
    label_logo1 = tk.Label(root, image=logo1_photo, bg="#f4faff")
    label_logo1.grid(row=0, column=0, padx=20, pady=10, sticky="nw", rowspan=2)

# โลโก้ 2
logo2_photo = load_image("hellow.png", (150, 150))
if logo2_photo:
    label_logo2 = tk.Label(root, image=logo2_photo, bg="#f4faff")
    label_logo2.grid(row=0, column=1, columnspan=2, pady=20, sticky="n")

# หัวข้อ
label_title = tk.Label(root, text="กรุณากรอกข้อมูลที่ต้องการค้นหา", font=("TH Sarabun", 28, "bold"), fg="#007bff", bg="#f4faff")
label_title.grid(row=1, column=0, columnspan=3, pady=10)

# ช่องกรอกคำค้นหา
entry_query = tk.Entry(root, font=("TH Sarabun", 20), bd=2, relief="solid", bg="#ffffff")
entry_query.grid(row=2, column=0, columnspan=3, pady=10, padx=20, sticky="ew")

# ปุ่มค้นหา
btn_search = tk.Button(root, text="ค้นหา", command=search_data, font=("TH Sarabun", 20), bg="#007bff", fg="white", relief="flat")
btn_search.grid(row=3, column=0, columnspan=3, pady=10)

# Text widget สำหรับผลการค้นหา
text_result = tk.Text(root, height=15, font=("TH Sarabun", 18), wrap="word", bd=2, relief="solid", bg="#ffffff")
text_result.grid(row=4, column=0, columnspan=3, pady=10, padx=20, sticky="nsew")

# Scrollbar
scrollbar = tk.Scrollbar(root, orient="vertical", command=text_result.yview)
scrollbar.grid(row=4, column=3, sticky="ns")
text_result.config(yscrollcommand=scrollbar.set)

# ปรับให้หน้าต่างยืดหยุ่น
root.grid_rowconfigure(4, weight=1)
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=1)
root.grid_columnconfigure(2, weight=1)

# เริ่มโปรแกรม
root.mainloop()